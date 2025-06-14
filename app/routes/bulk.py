from datetime import datetime
from typing import Optional
import csv
import io
import json

from fastapi import APIRouter, Request, Depends, Form, UploadFile, File
from fastapi.responses import RedirectResponse
from sqlalchemy import or_
import openpyxl
from sqlalchemy.orm import Session
import asyncssh

from app.models.models import (
    Device,
    VLAN,
    ConfigBackup,
    PortConfigTemplate,
    ImportLog,
    Site,
)
from app.utils.db_session import get_db
from app.utils.auth import require_role, get_user_site_ids
from app.utils.ssh import build_conn_kwargs, resolve_ssh_credential
from app.utils.device_detect import detect_ssh_platform
from app.utils.templates import templates
from app.utils.audit import log_audit
from app.utils.tags import get_or_create_tag, add_tag_to_device
from app.routes.devices import _format_ip

router = APIRouter(prefix="/bulk")


@router.get("/vlan-push")
async def bulk_vlan_push_form(
    request: Request,
    vlan_id: int | None = None,
    model_filter: str | None = None,
    db: Session = Depends(get_db),
    current_user=Depends(require_role("editor")),
):
    """Render form for pushing config to all devices in a VLAN."""
    site_ids = get_user_site_ids(db, current_user)
    vlan_ids = (
        db.query(Device.vlan_id)
        .filter(Device.site_id.in_(site_ids), Device.vlan_id.is_not(None))
        .distinct()
    )
    vlans = db.query(VLAN).filter(VLAN.id.in_(vlan_ids)).all()
    templates_db = db.query(PortConfigTemplate).all()

    device_count = None
    if vlan_id:
        q = db.query(Device).filter(
            Device.vlan_id == vlan_id, Device.site_id.in_(site_ids)
        )
        if model_filter:
            q = q.filter(Device.model.ilike(f"%{model_filter}%"))
        device_count = q.count()

    context = {
        "request": request,
        "vlans": vlans,
        "templates": templates_db,
        "selected_vlan": int(vlan_id) if vlan_id else None,
        "model_filter": model_filter or "",
        "device_count": device_count,
        "current_user": current_user,
    }
    return templates.TemplateResponse("bulk_vlan_push.html", context)


@router.post("/vlan-push")
async def bulk_vlan_push_action(
    request: Request,
    vlan_id: int = Form(...),
    template_id: Optional[int] = Form(None),
    config_text: str = Form(""),
    model_filter: str = Form(""),
    confirm: str | None = Form(None),
    db: Session = Depends(get_db),
    current_user=Depends(require_role("editor")),
):
    site_ids = get_user_site_ids(db, current_user)
    template = None
    if template_id:
        template = (
            db.query(PortConfigTemplate)
            .filter(PortConfigTemplate.id == template_id)
            .first()
        )
        if template:
            config_text = template.config_text
    q = db.query(Device).filter(Device.vlan_id == vlan_id, Device.site_id.in_(site_ids))
    if model_filter:
        q = q.filter(Device.model.ilike(f"%{model_filter}%"))
    devices = q.all()

    if not confirm:
        vlan = db.query(VLAN).filter(VLAN.id == vlan_id).first()
        context = {
            "request": request,
            "vlans": db.query(VLAN)
            .filter(
                VLAN.id.in_(
                    db.query(Device.vlan_id)
                    .filter(Device.site_id.in_(site_ids), Device.vlan_id.is_not(None))
                    .distinct()
                )
            )
            .all(),
            "templates": db.query(PortConfigTemplate).all(),
            "selected_vlan": vlan_id,
            "model_filter": model_filter,
            "config_text": config_text,
            "template_id": template_id,
            "device_count": len(devices),
            "confirm": True,
            "current_user": current_user,
        }
        return templates.TemplateResponse("bulk_vlan_push.html", context)

    for device in devices:
        if not device.ssh_credential and current_user.role != "superadmin":
            # skip if no credentials and user not superadmin
            continue
        cred, _ = resolve_ssh_credential(db, device, current_user)
        if not cred:
            continue
        conn_kwargs = build_conn_kwargs(cred)
        success = False
        try:
            async with asyncssh.connect(device.ip, **conn_kwargs) as conn:
                await detect_ssh_platform(db, device, conn, current_user)
                _, session = await conn.create_session(asyncssh.SSHClientProcess)
                for line in config_text.splitlines():
                    session.stdin.write(line + "\n")
                session.stdin.write("exit\n")
                await session.wait_closed()
                success = True
                device.last_seen = datetime.utcnow()
        except Exception:
            success = False
        backup = ConfigBackup(
            device_id=device.id,
            source="bulk_vlan_push",
            config_text=config_text,
            queued=not success,
            status="pushed" if success else "pending",
        )
        db.add(backup)
        db.commit()

    log_audit(
        db, current_user, "bulk_vlan_push", None, f"VLAN push to {len(devices)} devices"
    )
    return RedirectResponse(url="/tasks?message=Bulk+push+queued", status_code=302)


def _parse_upload(file: UploadFile) -> tuple[list[str], list[list[str]]]:
    """Return (columns, rows) parsed from CSV or XLSX upload."""
    content = file.file.read()
    file.file.close()
    if file.filename.lower().endswith(".xlsx"):
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        columns = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
        rows = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            rows.append([str(v) if v is not None else "" for v in r])
    else:
        reader = csv.reader(io.StringIO(content.decode()))
        data = list(reader)
        if not data:
            return [], []
        columns = [c.strip() for c in data[0]]
        rows = [list(r) for r in data[1:]]
    return columns, rows


def _map_rows(
    columns: list[str], rows: list[list[str]], mapping: dict[str, str]
) -> list[dict]:
    mapped: list[dict] = []
    for r in rows:
        data = {}
        for i, val in enumerate(r):
            field = mapping.get(str(i))
            if not field or field == "skip":
                continue
            data[field] = val.strip()
        mapped.append(data)
    return mapped


@router.get("/device-import")
async def device_import_form(
    request: Request,
    db: Session = Depends(get_db),
    current_user=Depends(require_role("editor")),
):
    sites = []
    if current_user.role == "superadmin":
        from app.models.models import Site

        sites = db.query(Site).all()
    context = {"request": request, "current_user": current_user, "sites": sites}
    return templates.TemplateResponse("device_import_upload.html", context)


@router.post("/device-import")
async def device_import_wizard(
    request: Request,
    import_file: UploadFile = File(None),
    file_data: str = Form(None),
    site_id: str = Form(None),
    confirm: str = Form(None),
    db: Session = Depends(get_db),
    current_user=Depends(require_role("editor")),
):
    if import_file:
        columns, rows = _parse_upload(import_file)
        data = json.dumps({"columns": columns, "rows": rows})
        fields = ["skip", "hostname", "ip", "device_type", "model", "vlan", "tags"]
        context = {
            "request": request,
            "columns": columns,
            "preview": rows[:5],
            "file_data": data,
            "fields": fields,
            "site_id": site_id or "",
            "current_user": current_user,
        }
        return templates.TemplateResponse("device_import_map.html", context)

    if file_data and not confirm:
        form = await request.form()
        raw = json.loads(file_data)
        mapping = {
            str(i): form.get(f"map_{i}", "skip") for i in range(len(raw["columns"]))
        }
        mapping_json = json.dumps(mapping)
        mapped_rows = _map_rows(raw["columns"], raw["rows"], mapping)
        duplicates = []
        for row in mapped_rows:
            ip = row.get("ip")
            host = row.get("hostname")
            if not ip and not host:
                continue
            existing = (
                db.query(Device)
                .filter(or_(Device.ip == ip, Device.hostname == host))
                .first()
            )
            if existing:
                duplicates.append({"row": row, "existing": existing})
        context = {
            "request": request,
            "duplicates": duplicates,
            "mapping_json": mapping_json,
            "file_data": file_data,
            "site_id": site_id or "",
            "current_user": current_user,
        }
        return templates.TemplateResponse("device_import_confirm.html", context)

    if confirm and file_data:
        mapping = json.loads(request.form.get("mapping_json"))
        raw = json.loads(file_data)
        mapped_rows = _map_rows(raw["columns"], raw["rows"], mapping)
        added = 0
        skipped = 0
        updated = 0
        errors = []
        user_site_ids = get_user_site_ids(db, current_user)
        chosen_site = int(site_id) if site_id else None
        if chosen_site is None and len(user_site_ids) == 1:
            chosen_site = user_site_ids[0]
        action = request.form.get("conflict_action", "skip")
        for row in mapped_rows:
            ip = row.get("ip")
            host = row.get("hostname")
            if not ip or not host:
                errors.append(f"Missing required fields in row {row}")
                continue
            ip = _format_ip(ip)
            existing = (
                db.query(Device)
                .filter(or_(Device.ip == ip, Device.hostname == host))
                .first()
            )
            if existing:
                if action == "skip":
                    skipped += 1
                    continue
                if action in {"overwrite", "merge"}:
                    if action == "overwrite":
                        existing.ip = ip
                        existing.hostname = host
                        if row.get("model"):
                            existing.model = row["model"]
                        if row.get("vlan"):
                            existing.vlan_id = row["vlan"]
                    else:
                        if not existing.model and row.get("model"):
                            existing.model = row["model"]
                        if not existing.vlan_id and row.get("vlan"):
                            existing.vlan_id = row["vlan"]
                    updated += 1
                    device = existing
                else:
                    skipped += 1
                    continue
            else:
                device = Device(
                    hostname=host,
                    ip=ip,
                    model=row.get("model"),
                    manufacturer=row.get("manufacturer") or "",
                    device_type_id=None,
                    vlan_id=row.get("vlan"),
                    site_id=chosen_site,
                    created_by_id=current_user.id,
                )
                db.add(device)
                added += 1
            tags_val = row.get("tags")
            if tags_val:
                names = [t.strip().lower() for t in tags_val.split(",") if t.strip()]
                for name in names:
                    tag = get_or_create_tag(db, name)
                    add_tag_to_device(db, device, tag, current_user)
        db.commit()
        db.add(
            ImportLog(
                user_id=current_user.id,
                file_name="upload",
                device_count=len(mapped_rows),
                site_id=chosen_site,
                notes="Imported",
                success=len(errors) == 0,
            )
        )
        db.commit()
        context = {
            "request": request,
            "added": added,
            "skipped": skipped,
            "updated": updated,
            "errors": errors,
            "current_user": current_user,
        }
        return templates.TemplateResponse("device_import_result.html", context)

    return RedirectResponse(url="/bulk/device-import", status_code=302)
